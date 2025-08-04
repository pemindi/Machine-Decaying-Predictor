import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

def create_excel_report(lower_results, upper_results, critical_threshold, session_id, model_manager, 
                       preprocessing_summaries=None, analysis_date=None, reports_folder=None):
    """Create comprehensive Excel report with detailed summaries"""
    wb = Workbook()
    wb.remove(wb.active)
    
    # Use current date if not provided
    if analysis_date is None:
        analysis_date = datetime.now()
    
    # Create summary sheet
    summary_sheet = wb.create_sheet("Executive_Summary")
    _create_executive_summary(summary_sheet, lower_results, upper_results, critical_threshold, preprocessing_summaries, analysis_date)
    
    # Create preprocessing summary sheet if available
    if preprocessing_summaries:
        preprocessing_sheet = wb.create_sheet("Preprocessing_Summary")
        _create_preprocessing_summary(preprocessing_sheet, preprocessing_summaries)
    
    # Create detailed sheets for each dataset
    _create_detailed_dataset_sheet(wb, "Lower_Dataset_Complete", lower_results, "Lower", critical_threshold, model_manager, analysis_date)
    _create_detailed_dataset_sheet(wb, "Upper_Dataset_Complete", upper_results, "Upper", critical_threshold, model_manager, analysis_date)
    
    # Create model comparison sheet
    comparison_sheet = wb.create_sheet("Model_Performance_Comparison")
    _create_model_comparison_sheet(comparison_sheet, lower_results, upper_results, model_manager)
    
    # Create critical dates overview
    critical_sheet = wb.create_sheet("Critical_Dates_Overview")
    _create_critical_dates_sheet(critical_sheet, lower_results, upper_results, critical_threshold, model_manager, analysis_date)
    
    # Save report
    filename = f"comprehensive_decay_report_{session_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(reports_folder, filename)
    wb.save(filepath)
    
    return filepath

def create_fallback_excel_report(lower_results, upper_results, session_id, analysis_date,
                                fallback_days, baseline_percentage, curve_fit_percentage, parameters, reports_folder):
    """Create comprehensive Excel report for fallback analysis"""
    wb = Workbook()
    wb.remove(wb.active)
    
    # Create executive summary sheet
    summary_sheet = wb.create_sheet("Executive_Summary")
    _create_fallback_executive_summary(summary_sheet, lower_results, upper_results, analysis_date,
                                     fallback_days, baseline_percentage, curve_fit_percentage, parameters)
    
    # Create detailed sheets for each dataset
    _create_fallback_detailed_dataset_sheet(wb, "Lower_Dataset_Complete", lower_results, "Lower", analysis_date)
    _create_fallback_detailed_dataset_sheet(wb, "Upper_Dataset_Complete", upper_results, "Upper", analysis_date)
    
    # Create model comparison sheet
    comparison_sheet = wb.create_sheet("Model_Performance_Comparison")
    _create_fallback_model_comparison_sheet(comparison_sheet, lower_results, upper_results)
    
    # Create failure predictions overview
    predictions_sheet = wb.create_sheet("Failure_Predictions_Overview")
    _create_fallback_predictions_sheet(predictions_sheet, lower_results, upper_results, analysis_date)
    
    # Create fallback analysis details
    fallback_sheet = wb.create_sheet("Fallback_Analysis_Details")
    _create_fallback_analysis_details_sheet(fallback_sheet, lower_results, upper_results)
    
    # Save report
    filename = f"comprehensive_fallback_report_{session_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(reports_folder, filename)
    wb.save(filepath)
    
    return filepath

# Helper functions for incremental reports
def _create_preprocessing_summary(sheet, preprocessing_summaries):
    """Create preprocessing summary sheet"""
    # Title
    sheet['A1'] = "Data Preprocessing Summary"
    sheet['A1'].font = Font(size=18, bold=True)
    sheet['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    sheet.merge_cells('A1:H1')
    
    row = 3
    
    # Headers
    headers = ['Dataset', 'Original Rows', 'Processed Rows', 'Rows Added', 'Date Range', 'Oversampling Method']
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    row += 1
    
    # Data rows
    for summary in preprocessing_summaries:
        data_row = [
            summary['dataset_type'].title(),
            summary['original_rows'],
            summary['processed_rows'],
            summary['rows_added'],
            summary['date_range'],
            summary['oversampling_method']
        ]
        
        for col, value in enumerate(data_row, 1):
            cell = sheet.cell(row=row, column=col, value=value)
            
            # Highlight rows added
            if col == 4 and isinstance(value, int) and value > 0:
                cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                cell.font = Font(bold=True)
        
        row += 1
    
    # Auto-adjust column widths
    column_widths = {'A': 12, 'B': 15, 'C': 15, 'D': 12, 'E': 30, 'F': 20}
    for col_letter, width in column_widths.items():
        try:
            sheet.column_dimensions[col_letter].width = width
        except:
            pass

def _create_executive_summary(sheet, lower_results, upper_results, critical_threshold, preprocessing_summaries=None, analysis_date=None):
    """Create executive summary sheet"""
    if analysis_date is None:
        analysis_date = datetime.now()
    
    # Title
    sheet['A1'] = "Machine Decay Prediction - Executive Summary"
    sheet['A1'].font = Font(size=18, bold=True)
    sheet['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    sheet.merge_cells('A1:H1')
    
    # Analysis details
    row = 3
    sheet[f'A{row}'] = f"Analysis Date: {analysis_date.strftime('%Y-%m-%d %H:%M:%S')}"
    sheet[f'A{row}'].font = Font(bold=True)
    row += 1
    sheet[f'A{row}'] = f"Critical Threshold: {critical_threshold}"
    sheet[f'A{row}'].font = Font(bold=True)
    row += 1
    
    # Preprocessing summary if available
    if preprocessing_summaries:
        sheet[f'A{row}'] = "Data Preprocessing Applied: ✓"
        sheet[f'A{row}'].font = Font(bold=True, color="00AA00")
        row += 1
        for summary in preprocessing_summaries:
            sheet[f'A{row}'] = f"  • {summary['dataset_type'].title()}: {summary['original_rows']} → {summary['processed_rows']} rows ({summary['oversampling_method']})"
            row += 1
    else:
        sheet[f'A{row}'] = "Data Preprocessing: Disabled (Minimal filtering applied)"
        sheet[f'A{row}'].font = Font(bold=True, color="FF6600")
        row += 1
    
    row += 1
    
    # Key findings
    sheet[f'A{row}'] = "Key Findings:"
    sheet[f'A{row}'].font = Font(size=14, bold=True)
    row += 1
    
    # Count total batches and critical predictions
    total_batches = 0
    critical_predictions = 0
    
    for dataset_name, results in [("Lower", lower_results), ("Upper", upper_results)]:
        for direction, direction_results in results.items():
            if direction_results:
                total_batches += len(direction_results)
                for batch_result in direction_results:
                    for critical_date in batch_result['critical_dates'].values():
                        if critical_date:
                            critical_predictions += 1
    
    sheet[f'A{row}'] = f"• Total Incremental Batches Analyzed: {total_batches}"
    row += 1
    sheet[f'A{row}'] = f"• Critical Date Predictions Found: {critical_predictions}"
    row += 1
    sheet[f'A{row}'] = f"• Datasets Processed: Lower and Upper bounds"
    row += 1
    sheet[f'A{row}'] = f"• Models Applied: 6 different regression models per batch"

def _create_detailed_dataset_sheet(wb, sheet_name, results, dataset_name, critical_threshold, model_manager, analysis_date=None):
    """Create detailed sheet for a dataset with all batch information"""
    if analysis_date is None:
        analysis_date = datetime.now()
    
    sheet = wb.create_sheet(sheet_name)
    
    # Title
    sheet['A1'] = f"{dataset_name} Dataset - Complete Batch Analysis"
    sheet['A1'].font = Font(size=16, bold=True)
    sheet['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    sheet.merge_cells('A1:I1')
    
    row = 3
    
    for direction, direction_results in results.items():
        if not direction_results:
            continue
            
        # Direction header
        sheet[f'A{row}'] = f"{direction.upper()} DIRECTION ANALYSIS"
        sheet[f'A{row}'].font = Font(size=14, bold=True)
        sheet[f'A{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        sheet.merge_cells(f'A{row}:I{row}')
        row += 2
        
        # Headers
        headers = ['Batch Size', 'Date Range', 'Model', 'R² Value', 'RMSE Value', 'Critical Date', 'Days from Today', 'Status', 'Model Rank']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        row += 1
        
        # Data rows
        for batch_result in direction_results:
            batch_models = []
            
            # Collect all model results for this batch
            for model_type, model_result in batch_result['model_results'].items():
                if model_result:
                    critical_date = batch_result['critical_dates'][model_type]
                    
                    batch_models.append({
                        'model_type': model_type,
                        'model_result': model_result,
                        'critical_date': critical_date,
                        'batch_size': batch_result['batch_size'],
                        'date_range': batch_result['date_range']
                    })
            
            # Sort models by R² value (best first)
            batch_models.sort(key=lambda x: x['model_result']['r2'], reverse=True)
            
            # Add data for each model
            for rank, model_data in enumerate(batch_models, 1):
                model_result = model_data['model_result']
                critical_date = model_data['critical_date']
                
                # Calculate days from today using the analysis date
                if critical_date:
                    days_diff = (critical_date - analysis_date).days
                    days_from_today = f"{days_diff}"
                    critical_date_str = critical_date.strftime('%Y-%m-%d %H:%M')
                    
                    # Determine status and color
                    if days_diff < 0:
                        status = "OVERDUE"
                        fill_color = "FF0000"  # Red
                    elif days_diff < 30:
                        status = "CRITICAL"
                        fill_color = "FF0000"  # Red
                    elif days_diff < 90:
                        status = "WARNING"
                        fill_color = "FFFF00"  # Yellow
                    else:
                        status = "ADVISORY"
                        fill_color = "FFFF00"  # Yellow
                else:
                    critical_date_str = "None"
                    days_from_today = "N/A"
                    status = "NO PREDICTION"
                    fill_color = "F0F0F0"  # Gray
                
                # Model name
                model_name = model_manager.model_names.get(model_data['model_type'], model_data['model_type'])
                
                data_row = [
                    model_data['batch_size'],
                    model_data['date_range'],
                    model_name,
                    round(model_result['r2'], 6),
                    round(model_result['rmse'], 4),
                    critical_date_str,
                    days_from_today,
                    status,
                    f"#{rank}"
                ]
                
                for col, value in enumerate(data_row, 1):
                    cell = sheet.cell(row=row, column=col, value=value)
                    
                    # Apply formatting
                    if col == 4:  # R² column
                        cell.number_format = '0.000000'
                    elif col == 5:  # RMSE column
                        cell.number_format = '0.0000'
                    elif col == 6 and critical_date:  # Critical date column
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                        if fill_color == "FF0000":  # Red text for red background
                            cell.font = Font(color="FFFFFF", bold=True)
                        elif fill_color == "FFFF00":  # Black text for yellow background
                            cell.font = Font(color="000000", bold=True)
                    elif col == 8:  # Status column
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                        if fill_color == "FF0000":
                            cell.font = Font(color="FFFFFF", bold=True)
                        elif fill_color == "FFFF00":
                            cell.font = Font(color="000000", bold=True)
                    elif col == 9:  # Rank column
                        if rank == 1:
                            cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                            cell.font = Font(bold=True)
                
                row += 1
            
            # Add separator between batches
            row += 1
        
        # Add separator between directions
        row += 2
    
    # Auto-adjust column widths
    column_widths = {
        'A': 12, 'B': 25, 'C': 20, 'D': 12, 'E': 12, 
        'F': 20, 'G': 15, 'H': 15, 'I': 12
    }
    for col_letter, width in column_widths.items():
        try:
            sheet.column_dimensions[col_letter].width = width
        except:
            pass

def _create_model_comparison_sheet(sheet, lower_results, upper_results, model_manager):
    """Create model performance comparison sheet"""
    sheet['A1'] = "Model Performance Comparison Across All Batches"
    sheet['A1'].font = Font(size=16, bold=True)
    sheet['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    sheet.merge_cells('A1:J1')
    
    row = 3
    
    # Headers
    headers = ['Dataset', 'Direction', 'Model', 'Avg R²', 'Best R²', 'Worst R²', 'Avg RMSE', 'Best RMSE', 'Worst RMSE', 'Success Rate']
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    row += 1
    
    # Calculate statistics for each model
    for dataset_name, results in [("Lower", lower_results), ("Upper", upper_results)]:
        for direction, direction_results in results.items():
            if not direction_results:
                continue
            
            # Collect all results by model type
            model_stats = {}
            
            for batch_result in direction_results:
                for model_type, model_result in batch_result['model_results'].items():
                    if model_type not in model_stats:
                        model_stats[model_type] = {'r2': [], 'rmse': [], 'total_attempts': 0}
                    
                    model_stats[model_type]['total_attempts'] += 1
                    
                    if model_result:
                        model_stats[model_type]['r2'].append(model_result['r2'])
                        model_stats[model_type]['rmse'].append(model_result['rmse'])
            
            # Write statistics
            for model_type, stats in model_stats.items():
                model_name = model_manager.model_names.get(model_type, model_type)
                
                if stats['r2']:
                    avg_r2 = sum(stats['r2']) / len(stats['r2'])
                    best_r2 = max(stats['r2'])
                    worst_r2 = min(stats['r2'])
                    avg_rmse = sum(stats['rmse']) / len(stats['rmse'])
                    best_rmse = min(stats['rmse'])
                    worst_rmse = max(stats['rmse'])
                    success_rate = len(stats['r2']) / stats['total_attempts']
                else:
                    avg_r2 = best_r2 = worst_r2 = 0
                    avg_rmse = best_rmse = worst_rmse = 0
                    success_rate = 0
                
                data_row = [
                    dataset_name,
                    direction.title(),
                    model_name,
                    round(avg_r2, 6),
                    round(best_r2, 6),
                    round(worst_r2, 6),
                    round(avg_rmse, 4),
                    round(best_rmse, 4),
                    round(worst_rmse, 4),
                    f"{success_rate:.1%}"
                ]
                
                for col, value in enumerate(data_row, 1):
                    cell = sheet.cell(row=row, column=col, value=value)
                    
                    # Highlight best performing models
                    if col in [4, 5, 6]:  # R² columns
                        if isinstance(value, (int, float)) and value > 0.9:
                            cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                        elif isinstance(value, (int, float)) and value > 0.8:
                            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                
                row += 1
    
    # Auto-adjust column widths
    column_widths = {
        'A': 12, 'B': 12, 'C': 20, 'D': 12, 'E': 12,
        'F': 12, 'G': 12, 'H': 12, 'I': 12, 'J': 15
    }
    for col_letter, width in column_widths.items():
        try:
            sheet.column_dimensions[col_letter].width = width
        except:
            pass

def _create_critical_dates_sheet(sheet, lower_results, upper_results, critical_threshold, model_manager, analysis_date=None):
    """Create critical dates overview sheet"""
    if analysis_date is None:
        analysis_date = datetime.now()
    
    sheet['A1'] = "Critical Dates Overview - All Predictions"
    sheet['A1'].font = Font(size=16, bold=True)
    sheet['A1'].fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    sheet.merge_cells('A1:H1')
    
    row = 3
    
    # Headers
    headers = ['Dataset', 'Direction', 'Batch Size', 'Model', 'Critical Date', 'Days from Today', 'Urgency Level', 'Recommended Action']
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    row += 1
    
    # Collect all critical dates
    all_critical_dates = []
    
    for dataset_name, results in [("Lower", lower_results), ("Upper", upper_results)]:
        for direction, direction_results in results.items():
            if direction_results:
                for batch_result in direction_results:
                    for model_type, critical_date in batch_result['critical_dates'].items():
                        if critical_date:
                            days_diff = (critical_date - analysis_date).days
                            
                            all_critical_dates.append({
                                'dataset': dataset_name,
                                'direction': direction,
                                'batch_size': batch_result['batch_size'],
                                'model': model_manager.model_names.get(model_type, model_type),
                                'critical_date': critical_date,
                                'days_diff': days_diff
                            })
    
    # Sort by urgency (days from today)
    all_critical_dates.sort(key=lambda x: x['days_diff'])
    
    # Add data rows
    for critical_data in all_critical_dates:
        days_diff = critical_data['days_diff']
        
        # Determine urgency and color
        if days_diff < 0:
            urgency = "OVERDUE"
            fill_color = "FF0000"  # Red
            action = "IMMEDIATE MAINTENANCE REQUIRED"
        elif days_diff < 30:
            urgency = "CRITICAL"
            fill_color = "FF0000"  # Red
            action = "Schedule maintenance within 7 days"
        elif days_diff < 90:
            urgency = "WARNING"
            fill_color = "FFFF00"  # Yellow
            action = "Plan maintenance within 2 weeks"
        else:
            urgency = "ADVISORY"
            fill_color = "FFFF00"  # Yellow
            action = "Monitor and plan ahead"
        
        data_row = [
            critical_data['dataset'],
            critical_data['direction'].title(),
            critical_data['batch_size'],
            critical_data['model'],
            critical_data['critical_date'].strftime('%Y-%m-%d %H:%M'),
            f"{days_diff} days",
            urgency,
            action
        ]
        
        for col, value in enumerate(data_row, 1):
            cell = sheet.cell(row=row, column=col, value=value)
            
            # Apply color coding to critical date and urgency columns
            if col in [5, 6, 7]:  # Critical date, days, and urgency columns
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                if fill_color == "FF0000":
                    cell.font = Font(color="FFFFFF", bold=True)
                else:
                    cell.font = Font(color="000000", bold=True)
        
        row += 1
    
    # Auto-adjust column widths
    column_widths = {
        'A': 12, 'B': 12, 'C': 12, 'D': 20, 'E': 20,
        'F': 15, 'G': 15, 'H': 40
    }
    for col_letter, width in column_widths.items():
        try:
            sheet.column_dimensions[col_letter].width = width
        except:
            pass

# Helper functions for fallback reports
def _create_fallback_executive_summary(sheet, lower_results, upper_results, analysis_date,
                                     fallback_days, baseline_percentage, curve_fit_percentage, parameters):
    """Create executive summary for fallback analysis"""
    # Title
    sheet['A1'] = "Fallback Days Analysis - Executive Summary"
    sheet['A1'].font = Font(size=18, bold=True)
    sheet['A1'].fill = PatternFill(start_color="FD7E14", end_color="FD7E14", fill_type="solid")
    sheet.merge_cells('A1:H1')
    
    # Analysis details
    row = 3
    sheet[f'A{row}'] = f"Analysis Date: {analysis_date.strftime('%Y-%m-%d %H:%M:%S')}"
    sheet[f'A{row}'].font = Font(bold=True)
    row += 1
    sheet[f'A{row}'] = f"Fallback Days: {fallback_days}"
    sheet[f'A{row}'].font = Font(bold=True)
    row += 1
    sheet[f'A{row}'] = f"Baseline Percentage: {baseline_percentage * 100}%"
    sheet[f'A{row}'].font = Font(bold=True)
    row += 1
    sheet[f'A{row}'] = f"Curve Fit Percentage: {curve_fit_percentage * 100}%"
    sheet[f'A{row}'].font = Font(bold=True)
    row += 1
    sheet[f'A{row}'] = f"Parameters Analyzed: {', '.join(parameters)}"
    sheet[f'A{row}'].font = Font(bold=True)
    row += 2
    
    # Key findings
    sheet[f'A{row}'] = "Key Findings:"
    sheet[f'A{row}'].font = Font(size=14, bold=True)
    row += 1
    
    total_analyses = len(lower_results) + len(upper_results)
    successful_predictions = 0
    fallback_predictions = 0
    
    for result in lower_results + upper_results:
        if result['failure_date']:
            successful_predictions += 1
        if result['fallback_results']['selected_fit'] and result['fallback_results']['selected_fit']['fail_date']:
            fallback_predictions += 1
    
    sheet[f'A{row}'] = f"• Total Parameter Analyses: {total_analyses}"
    row += 1
    sheet[f'A{row}'] = f"• Main Analysis Predictions: {successful_predictions}"
    row += 1
    sheet[f'A{row}'] = f"• Fallback Analysis Predictions: {fallback_predictions}"
    row += 1
    sheet[f'A{row}'] = f"• Lower Dataset Results: {len(lower_results)}"
    row += 1
    sheet[f'A{row}'] = f"• Upper Dataset Results: {len(upper_results)}"

def _create_fallback_detailed_dataset_sheet(wb, sheet_name, results, dataset_name, analysis_date):
    """Create detailed sheet for fallback dataset analysis"""
    sheet = wb.create_sheet(sheet_name)
    
    # Title
    sheet['A1'] = f"{dataset_name} Dataset - Fallback Analysis Results"
    sheet['A1'].font = Font(size=16, bold=True)
    sheet['A1'].fill = PatternFill(start_color="FD7E14", end_color="FD7E14", fill_type="solid")
    sheet.merge_cells('A1:L1')
    
    row = 3
    
    # Headers
    headers = ['Parameter', 'Model Used', 'R² Score', 'RMSE', 'Threshold', 'Main Failure Date', 
              'Fallback Model', 'Fallback R²', 'Fallback Failure Date', 'Days from Today', 'Status', 'Confidence']
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="FD7E14", end_color="FD7E14", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    row += 1
    
    # Data rows
    for result in results:
        fallback_fit = result['fallback_results']['selected_fit']
        
        # Determine best failure date (fallback takes priority if available)
        best_failure_date = None
        if fallback_fit and fallback_fit['fail_date']:
            best_failure_date = fallback_fit['fail_date']
        elif result['failure_date']:
            best_failure_date = result['failure_date']
        
        # Calculate days from today
        if best_failure_date:
            days_diff = (best_failure_date - analysis_date).days
            days_from_today = f"{days_diff}"
            
            # Determine status
            if days_diff < 0:
                status = "OVERDUE"
                fill_color = "FF0000"
            elif days_diff < 30:
                status = "CRITICAL"
                fill_color = "FF0000"
            elif days_diff < 90:
                status = "WARNING"
                fill_color = "FFFF00"
            else:
                status = "ADVISORY"
                fill_color = "FFFF00"
        else:
            days_from_today = "N/A"
            status = "NO PREDICTION"
            fill_color = "F0F0F0"
        
        # Confidence level based on R² scores
        main_r2 = result['r2']
        fallback_r2 = fallback_fit['r2'] if fallback_fit else 0
        best_r2 = max(main_r2, fallback_r2)
        
        if best_r2 > 0.9:
            confidence = "HIGH"
        elif best_r2 > 0.7:
            confidence = "MEDIUM"
        elif best_r2 > 0.5:
            confidence = "LOW"
        else:
            confidence = "VERY LOW"
        
        data_row = [
            result['param'],
            result['model'],
            round(result['r2'], 6),
            round(result['rmse'], 4),
            round(result['threshold'], 2),
            result['failure_date'].strftime('%Y-%m-%d %H:%M') if result['failure_date'] else "None",
            f"Best Fit (R²={fallback_r2:.3f})" if fallback_fit else "None",
            round(fallback_r2, 6) if fallback_fit else "N/A",
            fallback_fit['fail_date'].strftime('%Y-%m-%d %H:%M') if fallback_fit and fallback_fit['fail_date'] else "None",
            days_from_today,
            status,
            confidence
        ]
        
        for col, value in enumerate(data_row, 1):
            cell = sheet.cell(row=row, column=col, value=value)
            
            # Apply formatting
            if col in [3, 8]:  # R² columns
                if isinstance(value, (int, float)):
                    cell.number_format = '0.000000'
            elif col == 4:  # RMSE column
                if isinstance(value, (int, float)):
                    cell.number_format = '0.0000'
            elif col in [6, 9] and best_failure_date:  # Failure date columns
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                if fill_color == "FF0000":
                    cell.font = Font(color="FFFFFF", bold=True)
                elif fill_color == "FFFF00":
                    cell.font = Font(color="000000", bold=True)
            elif col == 11:  # Status column
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                if fill_color == "FF0000":
                    cell.font = Font(color="FFFFFF", bold=True)
                elif fill_color == "FFFF00":
                    cell.font = Font(color="000000", bold=True)
            elif col == 12:  # Confidence column
                if confidence == "HIGH":
                    cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                    cell.font = Font(bold=True)
                elif confidence == "MEDIUM":
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    cell.font = Font(bold=True)
        
        row += 1
    
    # Auto-adjust column widths
    column_widths = {
        'A': 15, 'B': 20, 'C': 12, 'D': 12, 'E': 12, 'F': 20,
        'G': 20, 'H': 12, 'I': 20, 'J': 15, 'K': 15, 'L': 12
    }
    for col_letter, width in column_widths.items():
        try:
            sheet.column_dimensions[col_letter].width = width
        except:
            pass

def _create_fallback_model_comparison_sheet(sheet, lower_results, upper_results):
    """Create model performance comparison for fallback analysis"""
    sheet['A1'] = "Fallback Analysis - Model Performance Comparison"
    sheet['A1'].font = Font(size=16, bold=True)
    sheet['A1'].fill = PatternFill(start_color="FD7E14", end_color="FD7E14", fill_type="solid")
    sheet.merge_cells('A1:J1')
    
    row = 3
    
    # Headers
    headers = ['Dataset', 'Parameter', 'Main Model', 'Main R²', 'Main RMSE', 
              'Fallback R²', 'Fallback RMSE', 'Best Method', 'Prediction Success', 'Confidence Level']
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="FD7E14", end_color="FD7E14", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    row += 1
    
    # Data rows
    for dataset_name, results in [("Lower", lower_results), ("Upper", upper_results)]:
        for result in results:
            fallback_fit = result['fallback_results']['selected_fit']
            
            main_r2 = result['r2']
            main_rmse = result['rmse']
            fallback_r2 = fallback_fit['r2'] if fallback_fit else 0
            fallback_rmse = fallback_fit['rmse'] if fallback_fit else 0
            
            # Determine best method
            if fallback_r2 > main_r2:
                best_method = "Fallback"
            elif main_r2 > fallback_r2:
                best_method = "Main"
            else:
                best_method = "Equal"
            
            # Prediction success
            has_main_prediction = result['failure_date'] is not None
            has_fallback_prediction = fallback_fit and fallback_fit['fail_date'] is not None
            
            if has_main_prediction and has_fallback_prediction:
                prediction_success = "Both Methods"
            elif has_main_prediction:
                prediction_success = "Main Only"
            elif has_fallback_prediction:
                prediction_success = "Fallback Only"
            else:
                prediction_success = "No Prediction"
            
            # Confidence level
            best_r2 = max(main_r2, fallback_r2)
            if best_r2 > 0.9:
                confidence = "HIGH"
            elif best_r2 > 0.7:
                confidence = "MEDIUM"
            elif best_r2 > 0.5:
                confidence = "LOW"
            else:
                confidence = "VERY LOW"
            
            data_row = [
                dataset_name,
                result['param'],
                result['model'],
                round(main_r2, 6),
                round(main_rmse, 4),
                round(fallback_r2, 6) if fallback_fit else "N/A",
                round(fallback_rmse, 4) if fallback_fit else "N/A",
                best_method,
                prediction_success,
                confidence
            ]
            
            for col, value in enumerate(data_row, 1):
                cell = sheet.cell(row=row, column=col, value=value)
                
                # Highlight best performing methods
                if col in [4, 6]:  # R² columns
                    if isinstance(value, (int, float)) and value > 0.9:
                        cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                    elif isinstance(value, (int, float)) and value > 0.7:
                        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                elif col == 8:  # Best method column
                    if value == "Fallback":
                        cell.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
                        cell.font = Font(bold=True)
                elif col == 10:  # Confidence column
                    if value == "HIGH":
                        cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                        cell.font = Font(bold=True)
            
            row += 1
    
    # Auto-adjust column widths
    column_widths = {
        'A': 12, 'B': 15, 'C': 20, 'D': 12, 'E': 12,
        'F': 12, 'G': 12, 'H': 15, 'I': 18, 'J': 15
    }
    for col_letter, width in column_widths.items():
        try:
            sheet.column_dimensions[col_letter].width = width
        except:
            pass

def _create_fallback_predictions_sheet(sheet, lower_results, upper_results, analysis_date):
    """Create failure predictions overview for fallback analysis"""
    sheet['A1'] = "Failure Predictions Overview - All Methods"
    sheet['A1'].font = Font(size=16, bold=True)
    sheet['A1'].fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    sheet.merge_cells('A1:I1')
    
    row = 3
    
    # Headers
    headers = ['Dataset', 'Parameter', 'Method', 'Model', 'Failure Date', 'Days from Today', 
              'Urgency Level', 'R² Score', 'Recommended Action']
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    row += 1
    
    # Collect all predictions
    all_predictions = []
    
    for dataset_name, results in [("Lower", lower_results), ("Upper", upper_results)]:
        for result in results:
            # Main analysis prediction
            if result['failure_date']:
                days_diff = (result['failure_date'] - analysis_date).days
                all_predictions.append({
                    'dataset': dataset_name,
                    'parameter': result['param'],
                    'method': 'Main Analysis',
                    'model': result['model'],
                    'failure_date': result['failure_date'],
                    'days_diff': days_diff,
                    'r2': result['r2']
                })
            
            # Fallback analysis prediction
            fallback_fit = result['fallback_results']['selected_fit']
            if fallback_fit and fallback_fit['fail_date']:
                days_diff = (fallback_fit['fail_date'] - analysis_date).days
                all_predictions.append({
                    'dataset': dataset_name,
                    'parameter': result['param'],
                    'method': 'Fallback Analysis',
                    'model': f"Best Fit (R²={fallback_fit['r2']:.3f})",
                    'failure_date': fallback_fit['fail_date'],
                    'days_diff': days_diff,
                    'r2': fallback_fit['r2']
                })
    
    # Sort by urgency (days from today)
    all_predictions.sort(key=lambda x: x['days_diff'])
    
    # Add data rows
    for pred in all_predictions:
        days_diff = pred['days_diff']
        
        # Determine urgency and action
        if days_diff < 0:
            urgency = "OVERDUE"
            fill_color = "FF0000"
            action = "IMMEDIATE MAINTENANCE REQUIRED"
        elif days_diff < 30:
            urgency = "CRITICAL"
            fill_color = "FF0000"
            action = "Schedule maintenance within 7 days"
        elif days_diff < 90:
            urgency = "WARNING"
            fill_color = "FFFF00"
            action = "Plan maintenance within 2 weeks"
        else:
            urgency = "ADVISORY"
            fill_color = "FFFF00"
            action = "Monitor and plan ahead"
        
        data_row = [
            pred['dataset'],
            pred['parameter'],
            pred['method'],
            pred['model'],
            pred['failure_date'].strftime('%Y-%m-%d %H:%M'),
            f"{days_diff} days",
            urgency,
            round(pred['r2'], 6),
            action
        ]
        
        for col, value in enumerate(data_row, 1):
            cell = sheet.cell(row=row, column=col, value=value)
            
            # Apply color coding
            if col in [5, 6, 7]:  # Date, days, and urgency columns
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                if fill_color == "FF0000":
                    cell.font = Font(color="FFFFFF", bold=True)
                else:
                    cell.font = Font(color="000000", bold=True)
            elif col == 8:  # R² column
                if isinstance(value, (int, float)):
                    cell.number_format = '0.000000'
                    if value > 0.9:
                        cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                        cell.font = Font(bold=True)
        
        row += 1
    
    # Auto-adjust column widths
    column_widths = {
        'A': 12, 'B': 15, 'C': 18, 'D': 25, 'E': 20,
        'F': 15, 'G': 15, 'H': 12, 'I': 40
    }
    for col_letter, width in column_widths.items():
        try:
            sheet.column_dimensions[col_letter].width = width
        except:
            pass

def _create_fallback_analysis_details_sheet(sheet, lower_results, upper_results):
    """Create detailed fallback analysis information"""
    sheet['A1'] = "Fallback Analysis - Technical Details"
    sheet['A1'].font = Font(size=16, bold=True)
    sheet['A1'].fill = PatternFill(start_color="FD7E14", end_color="FD7E14", fill_type="solid")
    sheet.merge_cells('A1:K1')
    
    row = 3
    
    # Headers
    headers = ['Dataset', 'Parameter', 'Baseline Mean', 'Baseline Std', 'Threshold', 
              'Fallback Days Used', 'Fallback R²', 'Fallback RMSE', 'CI Width', 'Data Points', 'Time Range']
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="FD7E14", end_color="FD7E14", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    row += 1
    
    # Data rows
    for dataset_name, results in [("Lower", lower_results), ("Upper", upper_results)]:
        for result in results:
            fallback_fit = result['fallback_results']['selected_fit']
            
            data_row = [
                dataset_name,
                result['param'],
                round(result['mean'], 4),
                round(result['std'], 4),
                round(result['threshold'], 2),
                fallback_fit['days_back'] if fallback_fit else "N/A",
                round(fallback_fit['r2'], 6) if fallback_fit else "N/A",
                round(fallback_fit['rmse'], 4) if fallback_fit else "N/A",
                round(result['mean_ci_width'], 4),
                len(result['df']),
                f"{result['df']['Datetime'].min().strftime('%Y-%m-%d')} to {result['df']['Datetime'].max().strftime('%Y-%m-%d')}"
            ]
            
            for col, value in enumerate(data_row, 1):
                cell = sheet.cell(row=row, column=col, value=value)
                
                # Format numeric columns
                if col in [3, 4, 5, 8, 9] and isinstance(value, (int, float)):
                    cell.number_format = '0.0000'
                elif col == 7 and isinstance(value, (int, float)):
                    cell.number_format = '0.000000'
            
            row += 1
    
    # Auto-adjust column widths
    column_widths = {
        'A': 12, 'B': 15, 'C': 15, 'D': 15, 'E': 12,
        'F': 18, 'G': 12, 'H': 12, 'I': 12, 'J': 12, 'K': 25
    }
    for col_letter, width in column_widths.items():
        try:
            sheet.column_dimensions[col_letter].width = width
        except:
            pass
